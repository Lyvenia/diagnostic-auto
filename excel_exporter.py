"""
Export Excel de la flotte via openpyxl.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

_BLUE_FILL   = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT = Font(color="4A9EFF", bold=True, size=10)
_BOLD        = Font(bold=True, size=10)
_NORMAL      = Font(size=10)
_ALT_FILL    = PatternFill("solid", fgColor="F7F9FC")
_THIN        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hdr(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = _BLUE_FILL
    cell.font   = _HEADER_FONT
    cell.border = _BORDER
    cell.alignment = _CENTER
    return cell


def _cell(ws, row, col, value, alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _ALT_FILL if alt else PatternFill()
    cell.font      = _NORMAL
    cell.border    = _BORDER
    cell.alignment = _LEFT
    return cell


# ── Main export ───────────────────────────────────────────────────────────────

def export_fleet_excel(vehicles: list) -> bytes:
    wb = Workbook()

    _sheet_fleet(wb, vehicles)
    _sheet_diagnostics(wb, vehicles)
    _sheet_repairs(wb, vehicles)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sheet_fleet(wb: Workbook, vehicles: list):
    ws = wb.active
    ws.title = "Flotte"

    headers = ["VIN", "Marque", "Modèle", "Année", "1er diagnostic",
               "Nbre diagnostics", "Dernier statut", "Nbre réparations", "Notes"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    for r, v in enumerate(vehicles, 2):
        alt = r % 2 == 0
        hist = v.get("historique", [])
        reps = v.get("reparations", [])
        _cell(ws, r, 1, v.get("vin", ""), alt)
        _cell(ws, r, 2, v.get("marque", ""), alt)
        _cell(ws, r, 3, v.get("modele", ""), alt)
        _cell(ws, r, 4, v.get("annee", ""), alt)
        prem = v.get("premier_diagnostic", "")
        try:
            prem = datetime.fromisoformat(prem).strftime("%d/%m/%Y")
        except Exception:
            pass
        _cell(ws, r, 5, prem, alt)
        _cell(ws, r, 6, len(hist), alt)
        _cell(ws, r, 7, v.get("statut_dernier_diagnostic", "OK"), alt)
        _cell(ws, r, 8, len(reps), alt)
        _cell(ws, r, 9, v.get("notes", "")[:200], alt)

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _sheet_diagnostics(wb: Workbook, vehicles: list):
    ws = wb.create_sheet("Diagnostics")
    headers = ["VIN", "Marque", "Année", "Date", "Kilométrage",
               "Codes DTC", "Nbre codes", "Statut", "Résumé IA"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    r = 2
    for v in vehicles:
        for entry in v.get("historique", []):
            alt = r % 2 == 0
            codes = entry.get("dtc_codes", [])
            resume = entry.get("analyse_ia", {}).get("resume", "")
            _cell(ws, r, 1, v.get("vin", ""), alt)
            _cell(ws, r, 2, v.get("marque", ""), alt)
            _cell(ws, r, 3, v.get("annee", ""), alt)
            _cell(ws, r, 4, entry.get("date_affichage", ""), alt)
            _cell(ws, r, 5, entry.get("kilometrage", 0), alt)
            _cell(ws, r, 6, ", ".join(codes) if codes else "—", alt)
            _cell(ws, r, 7, len(codes), alt)
            _cell(ws, r, 8, entry.get("statut", "OK"), alt)
            _cell(ws, r, 9, resume[:300] if resume else "", alt)
            r += 1

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _sheet_repairs(wb: Workbook, vehicles: list):
    ws = wb.create_sheet("Réparations")
    headers = ["VIN", "Marque", "Année", "Date", "Description",
               "Pièces", "Coût (€)", "Technicien"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    r = 2
    for v in vehicles:
        for rep in v.get("reparations", []):
            alt = r % 2 == 0
            _cell(ws, r, 1, v.get("vin", ""), alt)
            _cell(ws, r, 2, v.get("marque", ""), alt)
            _cell(ws, r, 3, v.get("annee", ""), alt)
            _cell(ws, r, 4, rep.get("date_affichage", rep.get("date", "")), alt)
            _cell(ws, r, 5, rep.get("description", ""), alt)
            _cell(ws, r, 6, rep.get("pieces", ""), alt)
            _cell(ws, r, 7, rep.get("cout", ""), alt)
            _cell(ws, r, 8, rep.get("technicien", ""), alt)
            r += 1

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value or "")
                max_len = max(max_len, min(len(v), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
