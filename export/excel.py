"""
Export Excel de la flotte via openpyxl — RODIA by Lyvenia.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Palette RODIA ─────────────────────────────────────────────────────────────
# Vert forêt foncé (header), crème rosé (accent), vert clair (alt row)

_GREEN_FILL  = PatternFill("solid", fgColor="0C180C")   # vert nuit — header bg
_ACCENT_FILL = PatternFill("solid", fgColor="1A2E1A")   # vert forêt — sous-header
_ALT_FILL    = PatternFill("solid", fgColor="F4F7F4")   # vert très clair — lignes alt
_ROSE_FILL   = PatternFill("solid", fgColor="F2C4B8")   # rose crème — urgences

_HEADER_FONT = Font(color="E8B4A4", bold=True, size=10)  # rose crème sur fond vert
_BOLD        = Font(bold=True, size=10)
_NORMAL      = Font(size=10)
_URGENT_FONT = Font(color="C0392B", bold=True, size=10)

_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hdr(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _GREEN_FILL
    cell.font      = _HEADER_FONT
    cell.border    = _BORDER
    cell.alignment = _CENTER
    return cell


def _cell(ws, row, col, value, alt=False, urgent=False):
    cell = ws.cell(row=row, column=col, value=value)
    if urgent:
        cell.fill = _ROSE_FILL
        cell.font = _URGENT_FONT
    else:
        cell.fill = _ALT_FILL if alt else PatternFill()
        cell.font = _NORMAL
    cell.border    = _BORDER
    cell.alignment = _LEFT
    return cell


# ── Main export ───────────────────────────────────────────────────────────────

def export_fleet_excel(vehicles: list) -> bytes:
    wb = Workbook()

    _sheet_fleet(wb, vehicles)
    _sheet_diagnostics(wb, vehicles)
    _sheet_repairs(wb, vehicles)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sheet_fleet(wb: Workbook, vehicles: list):
    ws = wb.active
    ws.title = "Flotte"

    headers = [
        "VIN", "Marque", "Modèle", "Motorisation", "Année",
        "1er diagnostic", "Nbre diagnostics",
        "Dernier statut", "Nbre réparations", "Notes"
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    for r, v in enumerate(vehicles, 2):
        alt  = r % 2 == 0
        hist = v.get("historique", [])
        reps = v.get("reparations", [])

        statut  = v.get("statut_dernier_diagnostic", "OK")
        urgent  = statut == "URGENT"

        prem = v.get("premier_diagnostic", "")
        try:
            prem = datetime.fromisoformat(prem).strftime("%d/%m/%Y")
        except Exception:
            pass

        motorisation = v.get("motorisation", "") or ""

        _cell(ws, r, 1,  v.get("vin", ""),     alt)
        _cell(ws, r, 2,  v.get("marque", ""),   alt)
        _cell(ws, r, 3,  v.get("modele", ""),   alt)
        _cell(ws, r, 4,  motorisation,           alt)
        _cell(ws, r, 5,  v.get("annee", ""),    alt)
        _cell(ws, r, 6,  prem,                  alt)
        _cell(ws, r, 7,  len(hist),             alt)
        _cell(ws, r, 8,  statut,                alt, urgent=urgent)
        _cell(ws, r, 9,  len(reps),             alt)
        _cell(ws, r, 10, v.get("notes", "")[:200], alt)

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _sheet_diagnostics(wb: Workbook, vehicles: list):
    ws = wb.create_sheet("Diagnostics")
    headers = [
        "VIN", "Marque", "Motorisation", "Année", "Date", "Kilométrage",
        "Codes DTC", "Nbre codes", "Statut", "Résumé IA"
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    r = 2
    for v in vehicles:
        motorisation = v.get("motorisation", "") or ""
        for entry in v.get("historique", []):
            alt    = r % 2 == 0
            codes  = entry.get("dtc_codes", [])
            resume = entry.get("analyse_ia", {}).get("resume", "")
            statut = entry.get("statut", "OK")
            urgent = statut == "URGENT"

            _cell(ws, r, 1,  v.get("vin", ""),             alt)
            _cell(ws, r, 2,  v.get("marque", ""),           alt)
            _cell(ws, r, 3,  motorisation,                  alt)
            _cell(ws, r, 4,  v.get("annee", ""),            alt)
            _cell(ws, r, 5,  entry.get("date_affichage", ""), alt)
            _cell(ws, r, 6,  entry.get("kilometrage", 0),   alt)
            _cell(ws, r, 7,  ", ".join(codes) if codes else "—", alt)
            _cell(ws, r, 8,  len(codes),                    alt)
            _cell(ws, r, 9,  statut,                        alt, urgent=urgent)
            _cell(ws, r, 10, resume[:300] if resume else "", alt)
            r += 1

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _sheet_repairs(wb: Workbook, vehicles: list):
    ws = wb.create_sheet("Réparations")
    headers = [
        "VIN", "Marque", "Motorisation", "Année",
        "Date", "Description", "Pièces", "Coût (€)", "Technicien"
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    r = 2
    for v in vehicles:
        motorisation = v.get("motorisation", "") or ""
        for rep in v.get("reparations", []):
            alt = r % 2 == 0
            _cell(ws, r, 1, v.get("vin", ""),    alt)
            _cell(ws, r, 2, v.get("marque", ""), alt)
            _cell(ws, r, 3, motorisation,         alt)
            _cell(ws, r, 4, v.get("annee", ""),  alt)
            _cell(ws, r, 5, rep.get("date_affichage", rep.get("date", "")), alt)
            _cell(ws, r, 6, rep.get("description", ""), alt)
            _cell(ws, r, 7, rep.get("pieces", ""),      alt)
            _cell(ws, r, 8, rep.get("cout", ""),        alt)
            _cell(ws, r, 9, rep.get("technicien", ""),  alt)
            r += 1

    _auto_width(ws)
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v       = str(cell.value or "")
                max_len = max(max_len, min(len(v), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
